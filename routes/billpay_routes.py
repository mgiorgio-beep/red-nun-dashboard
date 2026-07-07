"""
Bill Pay routes — AP aging, payment recording, vendor setup, check printing, CSV export.
Blueprint: billpay_bp at /api/billpay/*
"""

import csv
import io
import os
import logging
from datetime import datetime, date

from flask import Blueprint, jsonify, request, send_file, send_from_directory
from integrations.toast.data_store import get_connection
from routes.auth_routes import login_required, admin_required, admin_or_accountant_required

logger = logging.getLogger(__name__)

billpay_bp = Blueprint("billpay_bp", __name__)


# ─────────────────────────────────────────────
#  INVOICES (AP Outstanding)
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/invoices")
@admin_or_accountant_required
def get_billpay_invoices():
    """Return confirmed invoices with payment info, filterable."""
    vendor = request.args.get("vendor")
    status = request.args.get("payment_status")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    due_from = request.args.get("due_date_from")
    due_to = request.args.get("due_date_to")
    location = request.args.get("location")

    conn = get_connection()
    where = ["si.status = 'confirmed'",
             "si.vendor_name IN (SELECT vendor_name FROM vendor_bill_pay WHERE bill_pay_enabled = 1)"]
    params = []

    if vendor:
        # Include variant vendor names (e.g. "UniFirst" + "UniFirst Corporation")
        conn2 = get_connection()
        variant_names = conn2.execute(
            "SELECT DISTINCT vendor_name FROM scanned_invoices WHERE vendor_name IS NOT NULL"
        ).fetchall()
        conn2.close()

        import re as _re2
        def _norm_v(n):
            n = n.lower().strip()
            for s in [', inc.', ', inc', ' inc.', ' inc', ', llc', ' llc', ', corp', ' corp',
                      ' corporation', ' company', ' homegrown']:
                if n.endswith(s):
                    n = n[:-len(s)].strip()
            n = n.replace(' and ', ' & ')
            n = _re2.sub(r'[.,\-]+', ' ', n)
            n = _re2.sub(r'\s+', ' ', n).strip()
            n = n.replace('foodservice', 'food service')
            return n

        target_norm = _norm_v(vendor)
        matches = [vendor]
        for r in variant_names:
            vn = r["vendor_name"]
            if vn != vendor and _norm_v(vn) == target_norm:
                matches.append(vn)

        if len(matches) == 1:
            where.append("si.vendor_name = ?")
            params.append(vendor)
        else:
            placeholders = ",".join(["?"] * len(matches))
            where.append(f"si.vendor_name IN ({placeholders})")
            params.extend(matches)
    if date_from:
        where.append("si.invoice_date >= ?")
        params.append(date_from)
    if date_to:
        where.append("si.invoice_date <= ?")
        params.append(date_to)
    if due_from:
        where.append("si.due_date >= ?")
        params.append(due_from)
    if due_to:
        where.append("si.due_date <= ?")
        params.append(due_to)
    if location:
        where.append("si.location = ?")
        params.append(location)

    today = date.today().isoformat()

    if status == "unpaid":
        # Outstanding = unpaid positive-balance invoices PLUS unused credits.
        # Credits carry a negative balance; they appear as their own line items
        # and net down the amount due (mirrors the US Foods / vendor portal).
        # This clause is additive — it never hides a positive-balance bill.
        where.append(
            "((( si.payment_status = 'unpaid' OR si.payment_status IS NULL ) "
            "AND ( si.balance > 0 OR si.balance IS NULL )) "
            "OR COALESCE(si.balance, si.total, 0) < 0)"
        )
    elif status == "partial":
        where.append("si.payment_status = 'partial'")
    elif status == "pending_review":
        where.append("si.payment_status = 'pending_review'")
    elif status == "paid":
        # A fully-paid invoice has a zero balance with a positive total.
        # Exclude credits (negative balance/total) — they are "unused", not "paid".
        where.append("(si.payment_status = 'paid' OR (si.balance <= 0 AND COALESCE(si.total, 0) > 0))")
    elif status == "overdue":
        where.append("(si.payment_status != 'paid' OR si.payment_status IS NULL)")
        where.append("si.due_date < ?")
        params.append(today)
    else:
        # "All" — exclude fully paid zero-balance invoices (those belong in Payments tab)
        where.append("NOT (si.payment_status = 'paid' AND COALESCE(si.balance, si.total) <= 0)")

    where_sql = " AND ".join(where)

    rows = conn.execute(f"""
        SELECT si.id, si.vendor_name, si.invoice_number, si.invoice_date,
               si.due_date, si.total, si.subtotal, si.tax,
               COALESCE(si.amount_paid, 0) as amount_paid,
               COALESCE(si.balance, si.total) as balance,
               si.payment_status, si.location, si.category, si.payment_url
        FROM scanned_invoices si
        WHERE {where_sql}
        ORDER BY si.due_date ASC NULLS LAST, si.invoice_date ASC
    """, params).fetchall()

    invoices = []
    for r in rows:
        due = r["due_date"]
        days_overdue = 0
        if due:
            try:
                due_dt = datetime.strptime(due, "%Y-%m-%d").date()
                diff = (date.today() - due_dt).days
                days_overdue = max(0, diff)
            except ValueError:
                pass

        bal = r["balance"] if r["balance"] is not None else (r["total"] or 0)
        # Credit guard: a credit (negative total) must never carry a POSITIVE
        # balance — that sign flip inflated a 4-invoice Colonial selection by
        # 2x the credit ($1,986 instead of $1,676 on 2026-07-06). Use the
        # credit's own total until the underlying row is repaired.
        if (r["total"] or 0) < 0 and bal > 0:
            bal = r["total"]
        paid_amt = r["amount_paid"] or 0
        ps = r["payment_status"] or "unpaid"
        if bal <= 0 and (r["total"] or 0) > 0:
            ps = "paid"
        elif paid_amt > 0 and bal > 0:
            ps = "partial"
        elif days_overdue > 0 and ps != "paid":
            ps = "overdue"

        invoices.append({
            "id": r["id"],
            "vendor_name": r["vendor_name"],
            "invoice_number": r["invoice_number"],
            "invoice_date": r["invoice_date"],
            "due_date": r["due_date"],
            "total": r["total"],
            "amount_paid": paid_amt,
            "balance": bal,
            "payment_status": ps,
            "days_overdue": days_overdue,
            "location": r["location"],
            "category": r["category"],
            "payment_url": r["payment_url"],
        })

    conn.close()
    return jsonify({"invoices": invoices, "count": len(invoices)})


@billpay_bp.route("/api/billpay/invoices/export-csv")
@admin_or_accountant_required
def export_invoices_csv():
    """Export outstanding invoices as CSV."""
    conn = get_connection()
    rows = conn.execute("""
        SELECT vendor_name, invoice_number, invoice_date, due_date,
               total, COALESCE(amount_paid, 0) as amount_paid,
               COALESCE(balance, total) as balance, payment_status, location
        FROM scanned_invoices
        WHERE status = 'confirmed'
          AND vendor_name IN (SELECT vendor_name FROM vendor_bill_pay WHERE bill_pay_enabled = 1)
        ORDER BY due_date ASC NULLS LAST
    """).fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["Vendor", "Invoice #", "Invoice Date", "Due Date",
                     "Total", "Paid", "Balance", "Status", "Location"])
    for r in rows:
        writer.writerow([
            r["vendor_name"], r["invoice_number"], r["invoice_date"],
            r["due_date"], f"{r['total']:.2f}" if r["total"] else "",
            f"{r['amount_paid']:.2f}", f"{r['balance']:.2f}" if r["balance"] else "",
            r["payment_status"] or "unpaid", r["location"] or ""
        ])

    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode("utf-8")),
        mimetype="text/csv",
        as_attachment=True,
        download_name=f"ap_invoices_{date.today().isoformat()}.csv"
    )


# ─────────────────────────────────────────────
#  AP AGING SUMMARY
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/aging-summary")
@admin_or_accountant_required
def aging_summary():
    """Return AP aging buckets and per-vendor totals."""
    conn = get_connection()
    today = date.today().isoformat()
    location = request.args.get("location")

    sql = """
        SELECT id, vendor_name, due_date, total,
               COALESCE(amount_paid, 0) as amount_paid,
               COALESCE(balance, total) as balance,
               payment_status
        FROM scanned_invoices
        WHERE status = 'confirmed'
          AND (payment_status != 'paid' OR payment_status IS NULL)
          AND (balance > 0 OR balance IS NULL)
          AND vendor_name IN (SELECT vendor_name FROM vendor_bill_pay WHERE bill_pay_enabled = 1)
    """
    params = []
    if location:
        sql += " AND LOWER(location) = LOWER(?)"
        params.append(location)
    rows = conn.execute(sql, params).fetchall()
    conn.close()

    buckets = {
        "current": {"count": 0, "total": 0.0},
        "1_30": {"count": 0, "total": 0.0},
        "31_60": {"count": 0, "total": 0.0},
        "61_90": {"count": 0, "total": 0.0},
        "90_plus": {"count": 0, "total": 0.0},
    }
    vendor_totals = {}
    total_outstanding = 0.0

    for r in rows:
        bal = r["balance"] if r["balance"] is not None else (r["total"] or 0)
        # Credit guard — see api_billpay_invoices: a negative-total credit
        # must never count as positive outstanding.
        if (r["total"] or 0) < 0 and bal > 0:
            bal = r["total"]
        if bal <= 0:
            continue
        total_outstanding += bal

        vendor = r["vendor_name"] or "Unknown"
        if vendor not in vendor_totals:
            vendor_totals[vendor] = {"count": 0, "total": 0.0}
        vendor_totals[vendor]["count"] += 1
        vendor_totals[vendor]["total"] += bal

        due = r["due_date"]
        days_overdue = 0
        if due:
            try:
                due_dt = datetime.strptime(due, "%Y-%m-%d").date()
                days_overdue = (date.today() - due_dt).days
            except ValueError:
                pass

        if days_overdue <= 0:
            buckets["current"]["count"] += 1
            buckets["current"]["total"] += bal
        elif days_overdue <= 30:
            buckets["1_30"]["count"] += 1
            buckets["1_30"]["total"] += bal
        elif days_overdue <= 60:
            buckets["31_60"]["count"] += 1
            buckets["31_60"]["total"] += bal
        elif days_overdue <= 90:
            buckets["61_90"]["count"] += 1
            buckets["61_90"]["total"] += bal
        else:
            buckets["90_plus"]["count"] += 1
            buckets["90_plus"]["total"] += bal

    # Sort vendors by total descending
    sorted_vendors = sorted(vendor_totals.items(), key=lambda x: -x[1]["total"])

    return jsonify({
        "buckets": buckets,
        "total_outstanding": round(total_outstanding, 2),
        "vendor_totals": [{"vendor": v, **d} for v, d in sorted_vendors],
    })


# ─────────────────────────────────────────────
#  VENDOR SETUP
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/vendors")
@admin_or_accountant_required
def get_billpay_vendors():
    """Return all vendors merged from vendors table + scanned_invoices, with bill pay settings."""
    conn = get_connection()

    # Get all vendor names from both sources, deduplicating fuzzy matches
    vendor_names = set()

    # From vendors table (canonical names)
    rows = conn.execute("SELECT name FROM vendors").fetchall()
    for r in rows:
        if r["name"]:
            vendor_names.add(r["name"])

    # Build a normalized lookup to detect fuzzy duplicates
    import re as _re
    def _norm(n):
        n = n.lower().strip()
        # Strip common suffixes
        for s in [', inc.', ', inc', ' inc.', ' inc', ', llc', ' llc', ', corp', ' corp',
                  ' corporation', ' company', ' homegrown']:
            if n.endswith(s):
                n = n[:-len(s)].strip()
        # Normalize "and" <-> "&", punctuation, whitespace
        n = n.replace(' and ', ' & ')
        n = _re.sub(r'[.,\-]+', ' ', n)
        n = _re.sub(r'\s+', ' ', n).strip()
        # Remove "food service" vs "foodservice" difference
        n = n.replace('foodservice', 'food service')
        return n

    canonical_norms = {_norm(n): n for n in vendor_names}

    # From scanned_invoices (confirmed) — skip if fuzzy-matches a canonical vendor
    rows = conn.execute(
        "SELECT DISTINCT vendor_name FROM scanned_invoices WHERE status = 'confirmed' AND vendor_name IS NOT NULL"
    ).fetchall()
    for r in rows:
        inv_name = r["vendor_name"]
        normed = _norm(inv_name)
        if normed not in canonical_norms:
            vendor_names.add(inv_name)

    # Get bill pay settings
    bp_rows = conn.execute("SELECT * FROM vendor_bill_pay").fetchall()
    bp_map = {r["vendor_name"]: dict(r) for r in bp_rows}

    # Get outstanding totals per vendor (filtered by location if provided)
    location = request.args.get("location")
    outs_sql = """
        SELECT vendor_name, COUNT(*) as inv_count, SUM(COALESCE(balance, total)) as total_outstanding
        FROM scanned_invoices
        WHERE status = 'confirmed'
          AND (payment_status != 'paid' OR payment_status IS NULL)
          AND (balance > 0 OR balance IS NULL)
    """
    outs_params = []
    if location:
        outs_sql += " AND location = ?"
        outs_params.append(location)
    outs_sql += " GROUP BY vendor_name"
    outstanding = conn.execute(outs_sql, outs_params).fetchall()
    # Build outstanding map with normalized keys so variants merge (e.g. "UniFirst" + "UniFirst Corporation")
    outstanding_map = {}
    _norm_to_canonical = {_norm(n): n for n in vendor_names}
    for r in outstanding:
        raw_name = r["vendor_name"]
        normed = _norm(raw_name)
        # Map to canonical vendor name if possible, else use raw
        canonical = _norm_to_canonical.get(normed, raw_name)
        if canonical in outstanding_map:
            outstanding_map[canonical]["count"] += r["inv_count"]
            outstanding_map[canonical]["total"] += (r["total_outstanding"] or 0)
        else:
            outstanding_map[canonical] = {"count": r["inv_count"], "total": r["total_outstanding"] or 0}

    conn.close()

    vendors = []
    for name in sorted(vendor_names):
        bp = bp_map.get(name, {})
        outs = outstanding_map.get(name, {"count": 0, "total": 0})
        vendors.append({
            "vendor_name": name,
            "bill_pay_enabled": bool(bp.get("bill_pay_enabled", False)),
            "portal_pay_enabled": bool(bp.get("portal_pay_enabled", False)),
            "payment_recipient": bp.get("payment_recipient"),
            "remit_address_1": bp.get("remit_address_1"),
            "remit_address_2": bp.get("remit_address_2"),
            "remit_city": bp.get("remit_city"),
            "remit_state": bp.get("remit_state"),
            "remit_zip": bp.get("remit_zip"),
            "account_number": bp.get("account_number"),
            "contact_email": bp.get("contact_email"),
            "phone": bp.get("phone"),
            "payment_method": bp.get("payment_method", "check"),
            "payment_term_type": bp.get("payment_term_type", "not_specified"),
            "payment_term_days": bp.get("payment_term_days"),
            "payment_term_day_of_month": bp.get("payment_term_day_of_month"),
            "pay_lead_days": bp.get("pay_lead_days"),
            "auto_pay": bool(bp.get("auto_pay", False)),
            "notes": bp.get("notes"),
            "outstanding_count": outs["count"],
            "outstanding_total": round(outs["total"], 2),
        })

    return jsonify({"vendors": vendors})


@billpay_bp.route("/api/billpay/vendors/<path:vendor_name>", methods=["GET"])
@admin_or_accountant_required
def get_billpay_vendor(vendor_name):
    """Get full bill pay details for a vendor."""
    conn = get_connection()
    row = conn.execute("SELECT * FROM vendor_bill_pay WHERE vendor_name = ?", (vendor_name,)).fetchone()
    conn.close()

    if row:
        return jsonify(dict(row))
    return jsonify({
        "vendor_name": vendor_name,
        "bill_pay_enabled": False,
        "payment_method": "check",
        "payment_term_type": "not_specified",
    })


@billpay_bp.route("/api/billpay/vendors/<path:vendor_name>/sample-invoice")
@admin_required
def vendor_sample_invoice(vendor_name):
    """Serve the most recent invoice image/thumbnail for a vendor."""
    import os
    from flask import send_from_directory
    conn = get_connection()
    row = conn.execute("""
        SELECT id, image_path FROM scanned_invoices
        WHERE vendor_name = ? AND status = 'confirmed' AND image_path IS NOT NULL
        ORDER BY invoice_date DESC LIMIT 1
    """, (vendor_name,)).fetchone()
    conn.close()

    if not row or not row["image_path"]:
        return jsonify({"error": "No invoice image found"}), 404

    image_path = row["image_path"]
    # Always serve the original full-resolution file (PDF or image)
    directory = os.path.dirname(image_path)
    filename = os.path.basename(image_path)
    return send_from_directory(directory, filename)


@billpay_bp.route("/api/billpay/vendors/<path:vendor_name>", methods=["PUT"])
@admin_required
def update_billpay_vendor(vendor_name):
    """Create or update vendor bill pay settings."""
    data = request.get_json()
    conn = get_connection()

    # Case-insensitive match so saving with different capitalization updates the
    # existing record instead of creating a duplicate row (e.g. "...LLC" vs
    # "...Llc"). Duplicates here previously broke auto-pay: the invoice matched
    # one row while the auto_pay flag lived on the other.
    existing = conn.execute(
        "SELECT id FROM vendor_bill_pay WHERE vendor_name = ? COLLATE NOCASE", (vendor_name,)
    ).fetchone()

    fields = {
        "bill_pay_enabled": data.get("bill_pay_enabled", 1),
        "portal_pay_enabled": data.get("portal_pay_enabled", 0),
        "auto_pay": data.get("auto_pay", 0),
        "payment_recipient": data.get("payment_recipient"),
        "remit_address_1": data.get("remit_address_1"),
        "remit_address_2": data.get("remit_address_2"),
        "remit_city": data.get("remit_city"),
        "remit_state": data.get("remit_state"),
        "remit_zip": data.get("remit_zip"),
        "account_number": data.get("account_number"),
        "contact_email": data.get("contact_email"),
        "phone": data.get("phone"),
        "payment_method": data.get("payment_method", "check"),
        "payment_term_type": data.get("payment_term_type", "not_specified"),
        "payment_term_days": data.get("payment_term_days"),
        "payment_term_day_of_month": data.get("payment_term_day_of_month"),
        "pay_lead_days": data.get("pay_lead_days"),
        "auto_pay": data.get("auto_pay", 0),
        "send_confirmation_email": data.get("send_confirmation_email", 1),
        "notes": data.get("notes"),
        "updated_at": datetime.now().isoformat(),
    }

    if existing:
        sets = ", ".join(f"{k} = ?" for k in fields)
        conn.execute(
            f"UPDATE vendor_bill_pay SET {sets} WHERE id = ?",
            list(fields.values()) + [existing["id"]]
        )
    else:
        fields["vendor_name"] = vendor_name
        cols = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)
        conn.execute(
            f"INSERT INTO vendor_bill_pay ({cols}) VALUES ({placeholders})",
            list(fields.values())
        )

    conn.commit()
    conn.close()

    return jsonify({"status": "ok", "vendor_name": vendor_name})


# ─────────────────────────────────────────────
#  PAYMENTS
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/payments")
@admin_or_accountant_required
def get_payments():
    """Payment history with linked invoices."""
    vendor = request.args.get("vendor")
    status = request.args.get("status")
    method = request.args.get("method")
    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")

    conn = get_connection()
    where = ["1=1"]
    params = []

    if vendor:
        where.append("p.vendor_name = ?")
        params.append(vendor)
    if status:
        where.append("p.status = ?")
        params.append(status)
    if method:
        where.append("p.payment_method = ?")
        params.append(method)
    if date_from:
        where.append("p.payment_date >= ?")
        params.append(date_from)
    if date_to:
        where.append("p.payment_date <= ?")
        params.append(date_to)

    where_sql = " AND ".join(where)
    rows = conn.execute(f"""
        SELECT p.*
        FROM ap_payments p
        WHERE {where_sql}
        ORDER BY p.payment_date DESC, p.created_at DESC
    """, params).fetchall()

    payments = []
    for r in rows:
        # Get linked invoices
        links = conn.execute("""
            SELECT pi.amount_applied, si.invoice_number, si.vendor_name, si.total
            FROM ap_payment_invoices pi
            JOIN scanned_invoices si ON si.id = pi.invoice_id
            WHERE pi.payment_id = ?
        """, (r["id"],)).fetchall()

        payments.append({
            "id": r["id"],
            "vendor_name": r["vendor_name"],
            "payment_date": r["payment_date"],
            "amount": r["amount"],
            "payment_method": r["payment_method"],
            "check_number": r["check_number"],
            "reference_number": r["reference_number"],
            "memo": r["memo"],
            "status": r["status"],
            "created_at": r["created_at"],
            "invoices": [{"invoice_number": l["invoice_number"], "amount_applied": l["amount_applied"]} for l in links],
        })

    conn.close()
    return jsonify({"payments": payments, "count": len(payments)})


@billpay_bp.route("/api/billpay/payments", methods=["POST"])
@admin_required
def create_payment():
    """Create a payment and link to invoices."""
    data = request.get_json()
    vendor_name = data.get("vendor_name")
    payment_date = data.get("payment_date", date.today().isoformat())
    amount = data.get("amount", 0)
    payment_method = data.get("payment_method", "check")
    check_number = data.get("check_number")
    reference_number = data.get("reference_number")
    memo = data.get("memo")
    invoice_ids = data.get("invoice_ids", [])
    amounts_per_invoice = data.get("amounts_per_invoice", [])

    if not vendor_name or amount <= 0:
        return jsonify({"error": "vendor_name and positive amount required"}), 400

    conn = get_connection()
    cursor = conn.cursor()

    # Create payment record
    cursor.execute("""
        INSERT INTO ap_payments (vendor_name, payment_date, amount, payment_method,
            check_number, reference_number, memo, status)
        VALUES (?, ?, ?, ?, ?, ?, ?, 'pending')
    """, (vendor_name, payment_date, amount, payment_method,
          check_number, reference_number, memo))
    payment_id = cursor.lastrowid

    # Link to invoices and update balances
    for i, inv_id in enumerate(invoice_ids):
        applied = amounts_per_invoice[i] if i < len(amounts_per_invoice) else 0

        cursor.execute("""
            INSERT INTO ap_payment_invoices (payment_id, invoice_id, amount_applied)
            VALUES (?, ?, ?)
        """, (payment_id, inv_id, applied))

        # Update invoice balance
        cursor.execute("""
            UPDATE scanned_invoices
            SET amount_paid = COALESCE(amount_paid, 0) + ?,
                balance = COALESCE(balance, total) - ?,
                payment_status = CASE
                    WHEN COALESCE(balance, total) - ? <= 0 THEN 'paid'
                    ELSE 'partial'
                END,
                paid_date = CASE
                    WHEN COALESCE(balance, total) - ? <= 0 THEN ?
                    ELSE paid_date
                END
            WHERE id = ?
        """, (applied, applied, applied, applied, payment_date, inv_id))

    # ── Mirror into vendor_payments for centralized view ──
    # Reference prefix and source must match the payment method — previously
    # both were hardcoded for checks regardless of actual method, which made
    # ACH/cash/etc. payments appear as checks on the Payments page.
    method_norm = (payment_method or "check").strip().lower()
    if method_norm == "check":
        ref_prefix = "CHK"
        mirror_source = "check"
    elif method_norm == "ach":
        ref_prefix = "ACH"
        mirror_source = "external"
    elif method_norm == "credit_card":
        ref_prefix = "CC"
        mirror_source = "external"
    elif method_norm == "cash":
        ref_prefix = "CASH"
        mirror_source = "external"
    else:
        ref_prefix = method_norm.upper()[:6] or "PMT"
        mirror_source = "external"
    ref = (f"{ref_prefix}-{check_number}" if check_number
           else (reference_number if reference_number else f"{ref_prefix}-AP{payment_id}"))

    try:
        vp_cur = cursor.execute(
            """INSERT INTO vendor_payments
               (vendor, location, payment_date, payment_ref, payment_method,
                payment_total, check_number, memo, status, source, ap_payment_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'pending', ?, ?)""",
            (vendor_name, None, payment_date, ref, payment_method,
             amount, check_number, memo, mirror_source, payment_id),
        )
        vp_id = vp_cur.lastrowid
        # Mirror invoice links
        for i, inv_id in enumerate(invoice_ids):
            applied = amounts_per_invoice[i] if i < len(amounts_per_invoice) else 0
            inv_row = cursor.execute(
                "SELECT invoice_number, invoice_date, due_date FROM scanned_invoices WHERE id = ?",
                (inv_id,),
            ).fetchone()
            if inv_row:
                cursor.execute(
                    """INSERT INTO vendor_payment_invoices
                       (payment_id, invoice_number, invoice_date, due_date, amount_paid)
                       VALUES (?, ?, ?, ?, ?)""",
                    (vp_id, inv_row["invoice_number"], inv_row["invoice_date"],
                     inv_row["due_date"], applied),
                )
    except Exception as e:
        logger.warning(f"Mirror to vendor_payments failed for payment #{payment_id}: {e}")

    conn.commit()
    conn.close()

    logger.info(f"Payment #{payment_id} created: {vendor_name} ${amount:.2f} ({payment_method})")
    return jsonify({"status": "ok", "payment_id": payment_id}), 201


@billpay_bp.route("/api/billpay/payments/<int:payment_id>/void", methods=["PUT"])
@admin_required
def void_payment(payment_id):
    """Void a payment and reverse invoice balances."""
    conn = get_connection()
    cursor = conn.cursor()

    payment = cursor.execute("SELECT * FROM ap_payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        conn.close()
        return jsonify({"error": "Payment not found"}), 404
    if payment["status"] == "void":
        conn.close()
        return jsonify({"error": "Payment already voided"}), 400

    # Reverse linked invoice balances
    links = cursor.execute(
        "SELECT invoice_id, amount_applied FROM ap_payment_invoices WHERE payment_id = ?",
        (payment_id,)
    ).fetchall()

    for link in links:
        # Recalculate balance from scratch to avoid double-reverse bugs
        cursor.execute("""
            UPDATE scanned_invoices
            SET amount_paid = COALESCE((
                    SELECT SUM(api.amount_applied)
                    FROM ap_payment_invoices api
                    JOIN ap_payments ap ON ap.id = api.payment_id
                    WHERE api.invoice_id = scanned_invoices.id
                      AND ap.status != 'void'
                      AND ap.id != ?
                ), 0),
                balance = total - COALESCE((
                    SELECT SUM(api.amount_applied)
                    FROM ap_payment_invoices api
                    JOIN ap_payments ap ON ap.id = api.payment_id
                    WHERE api.invoice_id = scanned_invoices.id
                      AND ap.status != 'void'
                      AND ap.id != ?
                ), 0),
                payment_status = CASE
                    WHEN total - COALESCE((
                        SELECT SUM(api.amount_applied)
                        FROM ap_payment_invoices api
                        JOIN ap_payments ap ON ap.id = api.payment_id
                        WHERE api.invoice_id = scanned_invoices.id
                          AND ap.status != 'void'
                          AND ap.id != ?
                    ), 0) <= 0 THEN 'paid'
                    WHEN COALESCE((
                        SELECT SUM(api.amount_applied)
                        FROM ap_payment_invoices api
                        JOIN ap_payments ap ON ap.id = api.payment_id
                        WHERE api.invoice_id = scanned_invoices.id
                          AND ap.status != 'void'
                          AND ap.id != ?
                    ), 0) > 0 THEN 'partial'
                    ELSE 'unpaid'
                END
            WHERE id = ?
        """, (payment_id, payment_id, payment_id, payment_id, link["invoice_id"]))

    # Mark payment as void
    now = datetime.now().isoformat()
    cursor.execute(
        "UPDATE ap_payments SET status = 'void', updated_at = ? WHERE id = ?",
        (now, payment_id)
    )

    # ── Also void the mirrored vendor_payment ──
    try:
        cursor.execute(
            "UPDATE vendor_payments SET status = 'void', updated_at = ? WHERE ap_payment_id = ?",
            (now, payment_id),
        )
    except Exception:
        pass

    conn.commit()
    conn.close()

    logger.info(f"Payment #{payment_id} voided — reversed ${payment['amount']:.2f}")
    return jsonify({"status": "ok"})


@billpay_bp.route("/api/billpay/invoices/<int:invoice_id>/mark-paid-external", methods=["POST"])
@admin_or_accountant_required
def mark_invoice_paid_external(invoice_id):
    """Mark an invoice as paid outside the dashboard (QBO, manual ACH, etc.).

    Creates an ap_payments row with payment_method='external' and status='cleared',
    zeroes the invoice balance, and mirrors into vendor_payments. Intended for
    bills paid via the Pay ↗ link (QBO/Stripe/etc.) where the dashboard just
    needs to reflect that the bill is settled.
    """
    data = request.get_json(silent=True) or {}
    payment_date = data.get("payment_date") or date.today().isoformat()
    reference = (data.get("reference") or "").strip() or None
    memo = (data.get("memo") or "Paid externally").strip()
    # Accept payment_method so manual ACH payments label correctly on the
    # Payments page. Defaults to 'external' for backwards compatibility.
    payment_method = (data.get("payment_method") or "external").strip().lower()

    conn = get_connection()
    cursor = conn.cursor()

    inv = cursor.execute(
        "SELECT id, vendor_name, total, COALESCE(balance, total) AS balance, "
        "payment_status, location FROM scanned_invoices WHERE id = ?",
        (invoice_id,),
    ).fetchone()
    if not inv:
        conn.close()
        return jsonify({"error": "Invoice not found"}), 404
    if inv["payment_status"] == "paid":
        conn.close()
        return jsonify({"error": "Invoice already marked paid"}), 400

    applied = float(inv["balance"] or 0)
    # applied > 0  → real external payment: write ap_payments + vendor_payments
    # applied <= 0 → credit memo or $0 adjustment: just close out the invoice,
    #               no new money moved so skip the payment-record inserts
    payment_id = None
    if applied > 0:
        cursor.execute(
            """INSERT INTO ap_payments
               (vendor_name, payment_date, amount, payment_method,
                reference_number, memo, status)
               VALUES (?, ?, ?, ?, ?, ?, 'cleared')""",
            (inv["vendor_name"], payment_date, applied, payment_method, reference, memo),
        )
        payment_id = cursor.lastrowid

        cursor.execute(
            """INSERT INTO ap_payment_invoices (payment_id, invoice_id, amount_applied)
               VALUES (?, ?, ?)""",
            (payment_id, invoice_id, applied),
        )

    close_memo = memo if applied > 0 else (
        memo if memo != "Paid externally" else "Credit/adjustment resolved"
    )
    cursor.execute(
        """UPDATE scanned_invoices
           SET amount_paid = COALESCE(total, 0),
               balance = 0,
               payment_status = 'paid',
               paid_date = ?,
               payment_reference = COALESCE(?, payment_reference),
               notes = COALESCE(notes, '') || ' | ' || ?
           WHERE id = ?""",
        (payment_date, reference, f"resolved {payment_date}: {close_memo}", invoice_id),
    )

    if applied <= 0:
        conn.commit()
        conn.close()
        logger.info(f"Invoice #{invoice_id} ({inv['vendor_name']}) resolved — credit/zero balance, no ap_payments row (amount={applied:.2f})")
        return jsonify({"status": "ok", "payment_id": None, "amount": applied, "mode": "credit_resolved"})

    ref = reference or f"EXT-AP{payment_id}"
    try:
        vp_cur = cursor.execute(
            """INSERT INTO vendor_payments
               (vendor, location, payment_date, payment_ref, payment_method,
                payment_total, memo, status, source, ap_payment_id)
               VALUES (?, ?, ?, ?, ?, ?, ?, 'cleared', 'external', ?)""",
            (inv["vendor_name"], inv["location"], payment_date, ref,
             payment_method, applied, memo, payment_id),
        )
        vp_id = vp_cur.lastrowid
        inv_row = cursor.execute(
            "SELECT invoice_number, invoice_date, due_date FROM scanned_invoices WHERE id = ?",
            (invoice_id,),
        ).fetchone()
        if inv_row:
            cursor.execute(
                """INSERT INTO vendor_payment_invoices
                   (payment_id, invoice_number, invoice_date, due_date, amount_paid)
                   VALUES (?, ?, ?, ?, ?)""",
                (vp_id, inv_row["invoice_number"], inv_row["invoice_date"],
                 inv_row["due_date"], applied),
            )
    except Exception as e:
        logger.warning(f"Mirror to vendor_payments failed for external payment #{payment_id}: {e}")

    conn.commit()
    conn.close()

    logger.info(f"Invoice #{invoice_id} ({inv['vendor_name']}) marked paid externally — ap_payment #{payment_id} ${applied:.2f}")
    return jsonify({"status": "ok", "payment_id": payment_id, "amount": applied})


# ─────────────────────────────────────────────
#  PORTAL RECONCILIATION
# ─────────────────────────────────────────────

PORTAL_SNAPSHOT_DIR = os.path.expanduser("~rednun/vendor-scrapers/open_invoices")

# Map vendor_key (used in snapshot JSON filename prefix) to canonical vendor_name
# as stored in scanned_invoices.vendor_name.
PORTAL_VENDOR_MAP = {
    "pfg": "Performance Foodservice",
    "usfoods": "US Foods",
    "lknife": "L. Knife & Son, Inc.",
    "colonial": "Colonial Wholesale Beverage",
    "martignetti": "Martignetti Companies",
    "sg": "Southern Glazer's Beverage Company",
    "craft": "Craft Collective Inc",
}


@billpay_bp.route("/api/billpay/reconcile-portal", methods=["GET"])
@admin_or_accountant_required
def reconcile_portal():
    """Compare portal open-invoice snapshots to dashboard outstanding.

    Query params:
        vendor_key: e.g. 'pfg'. Required.

    Reads ~rednun/vendor-scrapers/open_invoices/<vendor_key>_<location>.json
    (one per location) and returns per-location diffs:
        - matched: in both portal and dashboard
        - portal_only: in portal, missing from dashboard (needs import)
        - dashboard_only: in dashboard, gone from portal (likely paid IRL — ghost)
        - amount_mismatch: in both but totals differ >$0.01
    """
    import glob
    vendor_key = (request.args.get("vendor_key") or "").strip().lower()
    if not vendor_key or vendor_key not in PORTAL_VENDOR_MAP:
        return jsonify({
            "error": "vendor_key required",
            "supported": list(PORTAL_VENDOR_MAP.keys()),
        }), 400

    vendor_name = PORTAL_VENDOR_MAP[vendor_key]

    snapshot_paths = sorted(glob.glob(os.path.join(
        PORTAL_SNAPSHOT_DIR, f"{vendor_key}_*.json"
    )))
    if not snapshot_paths:
        return jsonify({
            "error": f"No portal snapshots found for {vendor_key}",
            "expected_dir": PORTAL_SNAPSHOT_DIR,
        }), 404

    conn = get_connection()

    import json as _json
    results = []
    for path in snapshot_paths:
        try:
            with open(path) as f:
                snap = _json.load(f)
        except Exception as e:
            logger.warning(f"Failed to read portal snapshot {path}: {e}")
            continue

        location = snap.get("location")
        scraped_at = snap.get("scraped_at")
        portal_invs = snap.get("invoices", [])
        portal_by_num = {str(i["invoice_number"]).strip(): i for i in portal_invs if i.get("invoice_number")}

        db_rows = conn.execute(
            """SELECT id, invoice_number, invoice_date, total, balance, payment_status, image_path
               FROM scanned_invoices
               WHERE vendor_name = ? AND location = ?
                 AND status = 'confirmed'
                 AND (payment_status IS NULL OR payment_status != 'paid')""",
            (vendor_name, location),
        ).fetchall()
        db_by_num = {str(r["invoice_number"]).strip(): dict(r) for r in db_rows if r["invoice_number"]}

        matched, mismatches, portal_only, dashboard_only = [], [], [], []

        for num, portal_inv in portal_by_num.items():
            if num in db_by_num:
                db_inv = db_by_num[num]
                portal_amt = float(portal_inv.get("amount") or 0)
                db_amt = float(db_inv.get("total") or 0)
                if abs(portal_amt - db_amt) > 0.01:
                    mismatches.append({
                        "invoice_number": num,
                        "invoice_date": portal_inv.get("invoice_date"),
                        "portal_amount": portal_amt,
                        "dashboard_amount": db_amt,
                        "dashboard_id": db_inv["id"],
                    })
                else:
                    matched.append({
                        "invoice_number": num,
                        "invoice_date": portal_inv.get("invoice_date"),
                        "amount": portal_amt,
                        "dashboard_id": db_inv["id"],
                    })
            else:
                portal_only.append({
                    "invoice_number": num,
                    "invoice_date": portal_inv.get("invoice_date"),
                    "amount": float(portal_inv.get("amount") or 0),
                    "type": portal_inv.get("type"),
                })

        for num, db_inv in db_by_num.items():
            if num not in portal_by_num:
                dashboard_only.append({
                    "invoice_number": num,
                    "invoice_date": db_inv.get("invoice_date"),
                    "amount": float(db_inv.get("total") or 0),
                    "dashboard_id": db_inv["id"],
                })

        results.append({
            "location": location,
            "scraped_at": scraped_at,
            "portal_count": len(portal_invs),
            "dashboard_count": len(db_rows),
            "matched": matched,
            "portal_only": portal_only,
            "dashboard_only": dashboard_only,
            "amount_mismatch": mismatches,
        })

    conn.close()
    return jsonify({
        "vendor_key": vendor_key,
        "vendor_name": vendor_name,
        "locations": results,
    })


# ─────────────────────────────────────────────
#  CHECK CONFIG
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/check-config")
@admin_required
def get_check_config():
    """Return check printing config for a location."""
    location = request.args.get("location", "chatham")
    conn = get_connection()
    row = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not row:
        row = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    conn.close()
    if row:
        return jsonify(dict(row))
    return jsonify({})


@billpay_bp.route("/api/billpay/check-config", methods=["PUT"])
@admin_required
def update_check_config():
    """Save check config for a location."""
    data = request.get_json()
    location = data.get("location", "chatham")
    conn = get_connection()

    existing = conn.execute("SELECT id FROM check_config WHERE location = ?", (location,)).fetchone()
    fields = {
        "bank_name": data.get("bank_name"),
        "bank_address": data.get("bank_address"),
        "account_name": data.get("account_name"),
        "account_address_1": data.get("account_address_1"),
        "account_address_2": data.get("account_address_2"),
        "account_city_state_zip": data.get("account_city_state_zip"),
        "routing_number": data.get("routing_number"),
        "account_number": data.get("account_number"),
        "check_number_next": data.get("check_number_next"),
        "check_style": data.get("check_style", "top"),
        "offset_x": data.get("offset_x", 0),
        "offset_y": data.get("offset_y", 0),
        "signature_path": data.get("signature_path"),
        "micr_font_path": data.get("micr_font_path"),
        "location": location,
        "updated_at": datetime.now().isoformat(),
    }

    if existing:
        sets = ", ".join(f"{k} = ?" for k in fields)
        conn.execute(f"UPDATE check_config SET {sets} WHERE location = ?",
                     list(fields.values()) + [location])
    else:
        cols = ", ".join(fields.keys())
        placeholders = ", ".join("?" for _ in fields)
        conn.execute(f"INSERT INTO check_config ({cols}) VALUES ({placeholders})", list(fields.values()))

    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


@billpay_bp.route("/api/billpay/upload-signature", methods=["POST"])
@admin_required
def upload_signature():
    """Upload signature image for check printing."""
    if 'signature' not in request.files:
        return jsonify({"error": "No file provided"}), 400
    f = request.files['signature']
    if not f.filename:
        return jsonify({"error": "No file selected"}), 400

    sig_dir = "/opt/red-nun-dashboard/integrations/quickbooks/check_assets"
    os.makedirs(sig_dir, exist_ok=True)
    sig_path = os.path.join(sig_dir, "signature.png")

    # Save as PNG
    from PIL import Image
    img = Image.open(f).convert("RGBA")
    img.save(sig_path, "PNG")

    logger.info(f"Signature uploaded: {sig_path} ({img.size[0]}x{img.size[1]})")
    return jsonify({"status": "ok", "path": sig_path})


@billpay_bp.route("/api/billpay/signature-preview")
@admin_required
def signature_preview():
    """Serve the uploaded signature image."""
    sig_path = "/opt/red-nun-dashboard/integrations/quickbooks/check_assets/signature.png"
    if not os.path.exists(sig_path):
        return jsonify({"error": "No signature uploaded"}), 404
    return send_file(sig_path, mimetype="image/png")


@billpay_bp.route("/api/billpay/print-sample-check", methods=["POST"])
@admin_required
def print_sample_check():
    """Generate a sample check PDF using current config."""
    from check_printer import generate_check_pdf

    location = request.args.get("location")
    if not location and request.is_json:
        location = request.get_json(silent=True, force=True) or {}
        location = location.get("location", "chatham")
    if not location:
        location = "chatham"
    conn = get_connection()
    config = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not config:
        config = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    conn.close()

    if not config:
        return jsonify({"error": "No check config found"}), 400

    config = dict(config)
    sample_payment = {
        "amount": 1847.53,
        "vendor_name": "Sample Vendor Co.",
        "payment_date": datetime.now().strftime("%Y-%m-%d"),
        "memo": "SAMPLE — DO NOT DEPOSIT",
    }
    sample_vendor = {
        "payment_recipient": "Sample Vendor Co.",
        "remit_address_1": "123 Test Street",
        "remit_address_2": "",
        "remit_city": "Boston",
        "remit_state": "MA",
        "remit_zip": "02101",
    }
    sample_invoices = [
        {"invoice_number": "INV-0001", "invoice_date": "2026-03-01", "total": 1247.53, "amount_applied": 1247.53},
        {"invoice_number": "INV-0002", "invoice_date": "2026-03-08", "total": 600.00, "amount_applied": 600.00},
    ]

    check_num = config.get("check_number_next", 1001)
    output_path = f"/tmp/sample_check_{location}.pdf"
    generate_check_pdf(sample_payment, sample_invoices, config, sample_vendor,
                       check_number=f"SAMPLE", output_path=output_path)

    return send_file(output_path, mimetype="application/pdf",
                     download_name=f"sample_check_{location}.pdf")


# ─────────────────────────────────────────────
#  CHECK PRINTING
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/quick-check", methods=["POST"])
@admin_required
def quick_check():
    """Create a standalone (no-invoice) AP payment and immediately render its check PDF.

    Use for one-off / manual checks where there is no scanned invoice to pay
    against (reimbursements, ad-hoc vendor payments, etc.). The payment is
    inserted into ap_payments with no ap_payment_invoices links and mirrored
    into vendor_payments so it shows up on the Payments page.

    Request JSON:
      payee_name (str, required)      — name printed on the check
      amount (number, required)       — positive dollar amount
      location (str, optional)        — 'chatham' or 'dennis'; defaults to chatham
      payment_date (YYYY-MM-DD, opt)  — defaults to today
      memo (str, optional)            — printed on the check + stub
      check_number (str, optional)    — override; otherwise next from check_config
      address_1 (str, optional)       — envelope window line 1
      address_2 (str, optional)       — envelope window line 2 (or city/state/zip)
      gl_account_id (int, optional)   — chart-of-accounts row to code this against
    """
    from check_printer import generate_check_pdf

    data = request.get_json(silent=True) or {}
    payee = (data.get("payee_name") or "").strip()
    try:
        amount = float(data.get("amount") or 0)
    except (TypeError, ValueError):
        amount = 0
    if not payee:
        return jsonify({"error": "payee_name required"}), 400
    if amount <= 0:
        return jsonify({"error": "amount must be > 0"}), 400

    location = (data.get("location") or "chatham").strip().lower()
    payment_date = data.get("payment_date") or date.today().isoformat()
    memo = (data.get("memo") or "").strip() or None
    override_check_num = (data.get("check_number") or "").strip() or None
    addr1 = (data.get("address_1") or "").strip()
    addr2 = (data.get("address_2") or "").strip()
    try:
        gl_account_id = int(data.get("gl_account_id")) if data.get("gl_account_id") else None
    except (TypeError, ValueError):
        gl_account_id = None

    conn = get_connection()
    cursor = conn.cursor()

    # Resolve check config for the location (falls back to first row).
    config = conn.execute(
        "SELECT * FROM check_config WHERE location = ?", (location,)
    ).fetchone()
    if not config:
        config = conn.execute(
            "SELECT * FROM check_config ORDER BY id LIMIT 1"
        ).fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    # Pick the check number — override or next sequential.
    if override_check_num:
        check_num = override_check_num
    else:
        check_num = str(config["check_number_next"] or 1001)

    # Create the standalone payment row (status='printed' since we render now).
    cursor.execute(
        """INSERT INTO ap_payments (vendor_name, payment_date, amount, payment_method,
                check_number, memo, status, gl_account_id)
           VALUES (?, ?, ?, 'check', ?, ?, 'printed', ?)""",
        (payee, payment_date, amount, check_num, memo, gl_account_id),
    )
    payment_id = cursor.lastrowid

    # Bump the check sequence if we used it.
    if not override_check_num:
        cursor.execute(
            "UPDATE check_config SET check_number_next = ? WHERE location = ?",
            ((config["check_number_next"] or 1001) + 1, location),
        )

    # Mirror into vendor_payments so the Payments page sees it.
    try:
        cursor.execute(
            """INSERT INTO vendor_payments
               (vendor, location, payment_date, payment_ref, payment_method,
                payment_total, check_number, memo, status, source, ap_payment_id,
                gl_account_id)
               VALUES (?, ?, ?, ?, 'check', ?, ?, ?, 'printed', 'check', ?, ?)""",
            (payee, location, payment_date, f"CHK-{check_num}",
             amount, check_num, memo, payment_id, gl_account_id),
        )
    except Exception as e:
        logger.warning(f"Mirror to vendor_payments failed for quick-check #{payment_id}: {e}")

    conn.commit()

    payment_row = conn.execute(
        "SELECT * FROM ap_payments WHERE id = ?", (payment_id,)
    ).fetchone()

    # If we have an address from the modal, build a vendor_info-like dict so
    # `_build_payee_address` renders the envelope-window block. Otherwise pass
    # None and the check just shows the payee name.
    vendor_info = None
    if addr1 or addr2:
        vendor_info = {
            "payment_recipient": payee,
            "remit_address_1": addr1,
            "remit_address_2": addr2,
            "remit_city": "",
            "remit_state": "",
            "remit_zip": "",
        }
    else:
        # Try the bill-pay vendor table as a convenience for known vendors.
        vbp = conn.execute(
            "SELECT * FROM vendor_bill_pay WHERE vendor_name = ?", (payee,)
        ).fetchone()
        if vbp:
            vendor_info = dict(vbp)

    conn.close()

    output_path = f"/tmp/quick_check_{payment_id}_{check_num}.pdf"
    generate_check_pdf(
        payment=dict(payment_row),
        invoices=[],
        config=dict(config),
        vendor_info=vendor_info,
        check_number=check_num,
        output_path=output_path,
    )

    logger.info(
        f"Quick check generated: payment #{payment_id}, check #{check_num}, "
        f"${amount:,.2f} to {payee} ({location})"
    )

    return send_file(
        output_path,
        mimetype="application/pdf",
        download_name=f"check_{check_num}_{payee}.pdf",
        as_attachment=False,
    )


@billpay_bp.route("/api/billpay/payments/<int:payment_id>/print-check", methods=["GET", "POST"])
@admin_required
def print_check(payment_id):
    """Generate a PDF check for a payment."""
    from check_printer import generate_check_pdf

    conn = get_connection()
    payment = conn.execute("SELECT * FROM ap_payments WHERE id = ?", (payment_id,)).fetchone()
    if not payment:
        conn.close()
        return jsonify({"error": "Payment not found"}), 404

    # SAFETY (2026-05-28): refuse to print a check for a non-check payment method.
    # Previously this would happily generate a check PDF for ACH/credit-card/cash
    # payments AND write a check_number + status='printed' onto the ap_payment row,
    # silently corrupting non-check records. Now we hard-block.
    method = (payment["payment_method"] or "").strip().lower()
    if method and method != "check":
        conn.close()
        return jsonify({
            "error": (f"Payment #{payment_id} was recorded as '{method}', not 'check'. "
                      f"Cannot generate a check PDF for a non-check payment.")
        }), 400

    # Determine location from request or default to chatham
    location = request.args.get("location", "chatham")
    config = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not config:
        config = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    # Get linked invoices
    invoices = conn.execute("""
        SELECT pi.amount_applied, si.invoice_number, si.invoice_date, si.total
        FROM ap_payment_invoices pi
        JOIN scanned_invoices si ON si.id = pi.invoice_id
        WHERE pi.payment_id = ?
    """, (payment_id,)).fetchall()

    # Get vendor remittance address
    vendor_bp = conn.execute(
        "SELECT * FROM vendor_bill_pay WHERE vendor_name = ?",
        (payment["vendor_name"],)
    ).fetchone()

    # Assign check number if not set
    check_num = payment["check_number"]
    if not check_num:
        check_num = str(config["check_number_next"] or 1001)
        conn.execute("UPDATE ap_payments SET check_number = ?, status = 'printed' WHERE id = ?",
                      (check_num, payment_id))
        conn.execute("UPDATE check_config SET check_number_next = ? WHERE location = ?",
                      ((config["check_number_next"] or 1001) + 1, location))
        # ── Also update mirrored vendor_payment ──
        try:
            conn.execute(
                "UPDATE vendor_payments SET check_number = ?, status = 'printed', updated_at = ? WHERE ap_payment_id = ?",
                (check_num, datetime.now().isoformat(), payment_id),
            )
        except Exception:
            pass
        conn.commit()

    conn.close()

    output_path = f"/tmp/check_{payment_id}_{check_num}.pdf"
    generate_check_pdf(
        payment=dict(payment),
        invoices=[dict(i) for i in invoices],
        config=dict(config),
        vendor_info=dict(vendor_bp) if vendor_bp else None,
        check_number=check_num,
        output_path=output_path,
    )

    return send_file(output_path, mimetype="application/pdf",
                     download_name=f"check_{check_num}_{payment['vendor_name']}.pdf", as_attachment=False)


@billpay_bp.route("/api/billpay/payments/batch-print", methods=["POST"])
@admin_required
def batch_print_checks():
    """Generate multi-page PDF with one check per page."""
    from check_printer import generate_batch_checks_pdf

    data = request.get_json()
    payment_ids = data.get("payment_ids", [])
    if not payment_ids:
        return jsonify({"error": "No payment IDs provided"}), 400

    conn = get_connection()
    config = conn.execute("SELECT * FROM check_config WHERE id = 1").fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    payments_data = []
    next_check = config["check_number_next"] or 1001

    for pid in payment_ids:
        payment = conn.execute("SELECT * FROM ap_payments WHERE id = ?", (pid,)).fetchone()
        if not payment:
            continue

        invoices = conn.execute("""
            SELECT pi.amount_applied, si.invoice_number, si.invoice_date, si.total
            FROM ap_payment_invoices pi
            JOIN scanned_invoices si ON si.id = pi.invoice_id
            WHERE pi.payment_id = ?
        """, (pid,)).fetchall()

        vendor_bp = conn.execute(
            "SELECT * FROM vendor_bill_pay WHERE vendor_name = ?",
            (payment["vendor_name"],)
        ).fetchone()

        check_num = payment["check_number"]
        if not check_num:
            check_num = str(next_check)
            conn.execute("UPDATE ap_payments SET check_number = ?, status = 'printed' WHERE id = ?",
                          (check_num, pid))
            # Mirror to vendor_payments
            try:
                conn.execute(
                    "UPDATE vendor_payments SET check_number = ?, status = 'printed', updated_at = ? WHERE ap_payment_id = ?",
                    (check_num, datetime.now().isoformat(), pid),
                )
            except Exception:
                pass
            next_check += 1

        payments_data.append({
            "payment": dict(payment),
            "invoices": [dict(i) for i in invoices],
            "vendor_info": dict(vendor_bp) if vendor_bp else None,
            "check_number": check_num,
        })

    conn.execute("UPDATE check_config SET check_number_next = ? WHERE id = 1", (next_check,))
    conn.commit()
    conn.close()

    output_path = f"/tmp/batch_checks_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    generate_batch_checks_pdf(payments_data, dict(config), output_path)

    return send_file(output_path, mimetype="application/pdf",
                     download_name=f"checks_batch_{date.today().isoformat()}.pdf", as_attachment=False)


@billpay_bp.route("/api/billpay/print-calibration", methods=["POST"])
@admin_required
def print_calibration():
    """Generate calibration page for check alignment."""
    from check_printer import generate_calibration_page

    conn = get_connection()
    config = conn.execute("SELECT * FROM check_config WHERE id = 1").fetchone()
    conn.close()

    output_path = "/tmp/check_calibration.pdf"
    generate_calibration_page(dict(config) if config else {}, output_path)

    return send_file(output_path, mimetype="application/pdf",
                     download_name="check_calibration.pdf")


# ─────────────────────────────────────────────
#  PAYROLL CHECKS
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/payroll-checks")
@admin_or_accountant_required
def list_payroll_checks():
    """List payroll checks with optional filters."""
    location = request.args.get("location", "")
    conn = get_connection()
    where = []
    params = []
    if location:
        where.append("location = ?")
        params.append(location)
    where_sql = ("WHERE " + " AND ".join(where)) if where else ""
    rows = conn.execute(
        f"SELECT * FROM payroll_checks {where_sql} ORDER BY created_at DESC", params
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@billpay_bp.route("/api/billpay/payroll-checks", methods=["POST"])
@admin_required
def create_payroll_check():
    """Upload payroll PDF → OCR extract → assign check numbers → generate printable checks."""
    import anthropic
    import base64
    import json as json_mod
    from check_printer import generate_batch_payroll_checks_pdf

    pdf_file = request.files.get("pdf")
    if not pdf_file or not pdf_file.filename:
        return jsonify({"error": "PDF file required"}), 400

    location = request.form.get("location", "chatham")

    # Save the uploaded PDF
    payroll_dir = "/opt/red-nun-dashboard/payroll_checks"
    os.makedirs(payroll_dir, exist_ok=True)
    safe_name = f"payroll_{datetime.now().strftime('%Y%m%d%H%M%S')}_{pdf_file.filename}"
    safe_name = safe_name.replace(" ", "_")
    pdf_path = os.path.join(payroll_dir, safe_name)
    pdf_file.save(pdf_path)

    # OCR with Claude to extract check details
    records = []
    try:
        with open(pdf_path, "rb") as f:
            pdf_bytes = f.read()
        pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")

        client = anthropic.Anthropic()
        resp = client.messages.create(
            model=os.getenv("CLAUDE_MODEL", "claude-sonnet-4-6"),
            max_tokens=8192,
            messages=[{
                "role": "user",
                "content": [
                    {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
                    {"type": "text", "text": """Extract payroll check details from this PDF. Each page has one employee's pay stub/earnings statement.

IMPORTANT: Skip any employee whose net pay is $0.00 or negative — do NOT include them.

For EACH employee with net pay > 0, return a JSON array with objects containing:
- employee_name (string — full name)
- employee_address_1 (string — street address)
- employee_city (string)
- employee_state (string — 2-letter state code)
- employee_zip (string)
- pay_date (string, YYYY-MM-DD format — the check/pay date)
- gross_pay (number — current period gross earnings)
- net_pay (number — net pay / take-home amount)
- pay_period_start (string, YYYY-MM-DD)
- pay_period_end (string, YYYY-MM-DD)
- total_hours (number — total hours worked this period, or 0)
- deductions (object with keys: federal_tax, state_tax, fica_ss, fica_medicare — all numbers for CURRENT period. Include any other deductions in an "other" array of {label, amount} objects)
- ytd (object with keys: gross_pay, net_pay, federal_tax, state_tax, fica_ss, fica_medicare — all numbers for year-to-date totals)

Return ONLY the JSON array, no other text. Example:
[{"employee_name":"John Smith","employee_address_1":"123 Main St","employee_city":"Chatham","employee_state":"MA","employee_zip":"02633","pay_date":"2026-03-20","gross_pay":1500.00,"net_pay":1125.50,"total_hours":40.0,"pay_period_start":"2026-03-08","pay_period_end":"2026-03-14","deductions":{"federal_tax":150.00,"state_tax":75.00,"fica_ss":93.00,"fica_medicare":21.75,"other":[{"label":"MA Paid Family Leave","amount":5.37}]},"ytd":{"gross_pay":6000.00,"net_pay":4502.00,"federal_tax":600.00,"state_tax":300.00,"fica_ss":372.00,"fica_medicare":87.00}}]"""}
                ]
            }]
        )

        raw = resp.content[0].text.strip()
        if raw.startswith("```"):
            raw = raw.split("\n", 1)[1] if "\n" in raw else raw[3:]
            if raw.endswith("```"):
                raw = raw[:-3]
            raw = raw.strip()
        records = json_mod.loads(raw)
        logger.info(f"Payroll OCR extracted {len(records)} checks from {pdf_file.filename}")
    except Exception as e:
        logger.error(f"Payroll OCR failed: {e}")
        conn = get_connection()
        cursor = conn.execute("""INSERT INTO payroll_checks
            (employee_name, gross_pay, net_pay, location, pdf_path, updated_at)
            VALUES (?,0,0,?,?,datetime('now'))""",
            ("(OCR failed — edit manually)", location, pdf_path))
        conn.commit()
        check_id = cursor.lastrowid
        conn.close()
        return jsonify({"status": "ok", "id": check_id, "ocr_failed": True,
                        "message": f"PDF saved but OCR failed: {str(e)}"})

    # Get check config for this location
    conn = get_connection()
    config = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not config:
        config = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    config_dict = dict(config)
    next_check = config["check_number_next"] or 2011

    # Insert records and assign check numbers
    inserted_ids = []
    payroll_list = []  # for batch PDF generation
    for rec in records:
        check_num = str(next_check)
        deductions = rec.get("deductions", {})
        if isinstance(deductions, str):
            try:
                deductions = json_mod.loads(deductions)
            except (ValueError, TypeError):
                deductions = {}

        ytd = rec.get("ytd", {})
        total_hours = float(rec.get("total_hours", 0) or 0)

        cursor = conn.execute("""INSERT INTO payroll_checks
            (employee_name, employee_address_1, employee_city, employee_state, employee_zip,
             check_number, gross_pay, net_pay, deductions, total_hours, ytd_data,
             pay_period_start, pay_period_end,
             location, pdf_path, printed_at, status, updated_at)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,datetime('now'))""",
            (rec.get("employee_name", ""),
             rec.get("employee_address_1", ""),
             rec.get("employee_city", ""),
             rec.get("employee_state", ""),
             rec.get("employee_zip", ""),
             check_num,
             float(rec.get("gross_pay", 0) or 0),
             float(rec.get("net_pay", 0) or 0),
             json_mod.dumps(deductions),
             total_hours,
             json_mod.dumps(ytd),
             rec.get("pay_period_start"),
             rec.get("pay_period_end"),
             location, pdf_path,
             rec.get("pay_date"),
             "pending"))
        inserted_ids.append(cursor.lastrowid)

        # Build payroll dict for check printer
        payroll_list.append({
            "payroll": {
                "employee_name": rec.get("employee_name", ""),
                "employee_address_1": rec.get("employee_address_1", ""),
                "employee_city": rec.get("employee_city", ""),
                "employee_state": rec.get("employee_state", ""),
                "employee_zip": rec.get("employee_zip", ""),
                "gross_pay": float(rec.get("gross_pay", 0) or 0),
                "net_pay": float(rec.get("net_pay", 0) or 0),
                "total_hours": total_hours,
                "deductions": deductions,
                "ytd": ytd,
                "pay_period_start": rec.get("pay_period_start", ""),
                "pay_period_end": rec.get("pay_period_end", ""),
                "printed_at": rec.get("pay_date", ""),
                "memo": f"Payroll {rec.get('pay_period_start', '')} - {rec.get('pay_period_end', '')}",
            },
            "check_number": check_num,
        })
        next_check += 1

    # Update next check number
    conn.execute("UPDATE check_config SET check_number_next = ? WHERE location = ?",
                 (next_check, location))
    conn.commit()
    conn.close()

    # Generate printable checks PDF
    checks_pdf_path = os.path.join(payroll_dir,
        f"printable_checks_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
    generate_batch_payroll_checks_pdf(payroll_list, config_dict, checks_pdf_path)

    return jsonify({
        "status": "ok",
        "count": len(inserted_ids),
        "ids": inserted_ids,
        "checks_pdf": f"/api/billpay/payroll-checks/download-batch?path={checks_pdf_path}",
    })


@billpay_bp.route("/api/billpay/payroll-checks/<int:check_id>")
@admin_or_accountant_required
def get_payroll_check(check_id):
    """Get a single payroll check."""
    conn = get_connection()
    row = conn.execute("SELECT * FROM payroll_checks WHERE id = ?", (check_id,)).fetchone()
    conn.close()
    if not row:
        return jsonify({"error": "Not found"}), 404
    return jsonify(dict(row))


@billpay_bp.route("/api/billpay/payroll-checks/<int:check_id>", methods=["PUT"])
@admin_required
def update_payroll_check(check_id):
    """Update a payroll check (only if not yet printed)."""
    import json
    data = request.get_json()
    conn = get_connection()

    existing = conn.execute("SELECT * FROM payroll_checks WHERE id = ?", (check_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Not found"}), 404
    if existing["printed_at"]:
        conn.close()
        return jsonify({"error": "Cannot edit a printed check"}), 400

    gross = float(data.get("gross_pay", existing["gross_pay"]))
    deductions = data.get("deductions", existing["deductions"])
    if isinstance(deductions, str):
        deductions = json.loads(deductions)

    total_ded = sum(float(deductions.get(k, 0) or 0) for k in
                    ["federal_tax", "state_tax", "fica_ss", "fica_medicare"])
    for item in deductions.get("other", []):
        total_ded += float(item.get("amount", 0) or 0)
    net = gross - total_ded

    conn.execute("""UPDATE payroll_checks SET
        employee_name=?, employee_address_1=?, employee_address_2=?,
        employee_city=?, employee_state=?, employee_zip=?,
        gross_pay=?, net_pay=?, deductions=?, pay_period_start=?, pay_period_end=?,
        location=?, memo=?, updated_at=datetime('now')
        WHERE id=?""",
        (data.get("employee_name", existing["employee_name"]),
         data.get("employee_address_1", existing["employee_address_1"]),
         data.get("employee_address_2", existing["employee_address_2"]),
         data.get("employee_city", existing["employee_city"]),
         data.get("employee_state", existing["employee_state"]),
         data.get("employee_zip", existing["employee_zip"]),
         gross, net, json.dumps(deductions),
         data.get("pay_period_start", existing["pay_period_start"]),
         data.get("pay_period_end", existing["pay_period_end"]),
         data.get("location", existing["location"]),
         data.get("memo", existing["memo"]),
         check_id))

    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "net_pay": net})


@billpay_bp.route("/api/billpay/payroll-checks/<int:check_id>/print", methods=["POST"])
@admin_required
def print_payroll_check(check_id):
    """Print a payroll check — assigns check number from shared sequence."""
    from check_printer import generate_payroll_check_pdf

    conn = get_connection()
    payroll = conn.execute("SELECT * FROM payroll_checks WHERE id = ?", (check_id,)).fetchone()
    if not payroll:
        conn.close()
        return jsonify({"error": "Not found"}), 404
    if payroll["voided"]:
        conn.close()
        return jsonify({"error": "Cannot print a voided check"}), 400

    location = payroll["location"]
    config = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not config:
        config = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    # Assign check number if not already set
    check_num = payroll["check_number"]
    if not check_num:
        check_num = str(config["check_number_next"] or 1001)
        conn.execute("UPDATE payroll_checks SET check_number = ?, printed_at = datetime('now') WHERE id = ?",
                     (check_num, check_id))
        conn.execute("UPDATE check_config SET check_number_next = ? WHERE location = ?",
                     ((config["check_number_next"] or 1001) + 1, location))
        conn.commit()
    elif not payroll["printed_at"]:
        conn.execute("UPDATE payroll_checks SET printed_at = datetime('now') WHERE id = ?", (check_id,))
        conn.commit()

    conn.close()

    output_path = f"/tmp/payroll_check_{check_id}_{check_num}.pdf"
    generate_payroll_check_pdf(
        payroll=dict(payroll),
        config=dict(config),
        check_number=check_num,
        output_path=output_path,
    )

    return send_file(output_path, mimetype="application/pdf",
                     download_name=f"payroll_check_{check_num}_{payroll['employee_name']}.pdf", as_attachment=False)


@billpay_bp.route("/api/billpay/payroll-checks/<int:check_id>/pdf")
@admin_required
def view_payroll_pdf(check_id):
    """Serve the uploaded payroll source PDF."""
    conn = get_connection()
    row = conn.execute("SELECT pdf_path FROM payroll_checks WHERE id = ?", (check_id,)).fetchone()
    conn.close()
    if not row or not row["pdf_path"] or not os.path.exists(row["pdf_path"]):
        return jsonify({"error": "PDF not found"}), 404
    directory = os.path.dirname(row["pdf_path"])
    filename = os.path.basename(row["pdf_path"])
    return send_from_directory(directory, filename, mimetype="application/pdf")


@billpay_bp.route("/api/billpay/payroll-checks/<int:check_id>/generated-pdf")
@admin_required
def view_generated_check_pdf(check_id):
    """Serve the generated printable check PDF."""
    conn = get_connection()
    row = conn.execute("SELECT generated_pdf_path FROM payroll_checks WHERE id = ?", (check_id,)).fetchone()
    conn.close()
    if not row or not row["generated_pdf_path"] or not os.path.exists(row["generated_pdf_path"]):
        return jsonify({"error": "Generated check PDF not found"}), 404
    directory = os.path.dirname(row["generated_pdf_path"])
    filename = os.path.basename(row["generated_pdf_path"])
    return send_from_directory(directory, filename, mimetype="application/pdf")


@billpay_bp.route("/api/billpay/payroll-checks/download-batch")
@admin_required
def download_batch_payroll():
    """Download a generated batch payroll checks PDF."""
    path = request.args.get("path", "")
    if not path or not os.path.exists(path) or not path.startswith("/opt/red-nun-dashboard/payroll_checks/"):
        return jsonify({"error": "File not found"}), 404
    return send_file(path, mimetype="application/pdf",
                     download_name=f"payroll_checks_{date.today().isoformat()}.pdf")


@billpay_bp.route("/api/billpay/payroll-checks/print-all", methods=["GET", "POST"])
@admin_required
def print_all_payroll():
    """Generate printable checks PDF for all non-voided payroll checks in a location."""
    import json as json_mod
    from check_printer import generate_batch_payroll_checks_pdf

    location = request.args.get("location", "chatham")
    conn = get_connection()

    config = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not config:
        config = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    checks = conn.execute("""SELECT * FROM payroll_checks
        WHERE voided = 0 AND location = ?
        ORDER BY id""", (location,)).fetchall()

    if not checks:
        conn.close()
        return jsonify({"error": "No payroll checks to print"}), 400

    config_dict = dict(config)
    payroll_list = []
    for c in checks:
        ded = c["deductions"] or "{}"
        if isinstance(ded, str):
            try:
                ded = json_mod.loads(ded)
            except (ValueError, TypeError):
                ded = {}

        ytd_raw = c["ytd_data"] or "{}"
        if isinstance(ytd_raw, str):
            try:
                ytd_d = json_mod.loads(ytd_raw)
            except (ValueError, TypeError):
                ytd_d = {}
        else:
            ytd_d = ytd_raw or {}

        payroll_list.append({
            "payroll": {
                "employee_name": c["employee_name"] or "",
                "employee_address_1": c["employee_address_1"] or "",
                "employee_city": c["employee_city"] or "",
                "employee_state": c["employee_state"] or "",
                "employee_zip": c["employee_zip"] or "",
                "gross_pay": c["gross_pay"] or 0,
                "net_pay": c["net_pay"] or 0,
                "total_hours": float(c["total_hours"] or 0),
                "deductions": ded,
                "ytd": ytd_d,
                "pay_period_start": c["pay_period_start"] or "",
                "pay_period_end": c["pay_period_end"] or "",
                "printed_at": c["printed_at"] or "",
                "memo": c["memo"] or f"Payroll {c['pay_period_start'] or ''} - {c['pay_period_end'] or ''}",
            },
            "check_number": c["check_number"] or "",
        })

    # Generate and store the PDF
    payroll_dir = "/opt/red-nun-dashboard/payroll_checks"
    os.makedirs(payroll_dir, exist_ok=True)
    output_path = os.path.join(payroll_dir,
        f"printed_checks_{location}_{datetime.now().strftime('%Y%m%d%H%M%S')}.pdf")
    generate_batch_payroll_checks_pdf(payroll_list, config_dict, output_path)

    # Mark all as printed and store generated PDF path
    check_ids = [c["id"] for c in checks]
    for cid in check_ids:
        conn.execute("""UPDATE payroll_checks
            SET status = 'printed', generated_pdf_path = ?, updated_at = datetime('now')
            WHERE id = ?""", (output_path, cid))
    conn.commit()
    conn.close()

    return send_file(output_path, mimetype="application/pdf",
                     download_name=f"payroll_checks_{location}_{date.today().isoformat()}.pdf")


@billpay_bp.route("/api/billpay/payroll-checks/<int:check_id>/void", methods=["PUT"])
@admin_required
def void_payroll_check(check_id):
    """Void a payroll check."""
    conn = get_connection()
    existing = conn.execute("SELECT * FROM payroll_checks WHERE id = ?", (check_id,)).fetchone()
    if not existing:
        conn.close()
        return jsonify({"error": "Not found"}), 404

    conn.execute("UPDATE payroll_checks SET voided = 1, voided_at = datetime('now') WHERE id = ?",
                 (check_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
#  MANUAL CHECKS (dividends, misc, etc.)
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/manual-checks")
@admin_required
def list_manual_checks():
    """List manual checks with optional filters."""
    conn = get_connection()
    rows = conn.execute(
        "SELECT * FROM manual_checks ORDER BY created_at DESC"
    ).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])


@billpay_bp.route("/api/billpay/manual-checks", methods=["POST"])
@admin_required
def create_manual_check():
    """Create and optionally print a manual check."""
    data = request.get_json()
    payee = data.get("payee_name", "").strip()
    amount = float(data.get("amount", 0))
    if not payee or amount <= 0:
        return jsonify({"error": "Payee name and positive amount required"}), 400

    location = data.get("location", "chatham")
    conn = get_connection()

    cursor = conn.execute("""INSERT INTO manual_checks
        (payee_name, payee_address_1, payee_address_2,
         payee_city, payee_state, payee_zip,
         amount, memo, location, check_type, created_at, updated_at)
        VALUES (?,?,?,?,?,?,?,?,?,?,datetime('now'),datetime('now'))""",
        (payee, data.get("payee_address_1"), data.get("payee_address_2"),
         data.get("payee_city"), data.get("payee_state"), data.get("payee_zip"),
         amount, data.get("memo"), location, data.get("check_type", "manual")))

    check_id = cursor.lastrowid
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "id": check_id})


@billpay_bp.route("/api/billpay/manual-checks/<int:check_id>/print", methods=["GET", "POST"])
@admin_required
def print_manual_check(check_id):
    """Print a manual check — assigns check number from shared sequence."""
    from check_printer import generate_check_pdf

    conn = get_connection()
    mc = conn.execute("SELECT * FROM manual_checks WHERE id = ?", (check_id,)).fetchone()
    if not mc:
        conn.close()
        return jsonify({"error": "Not found"}), 404
    if mc["voided"]:
        conn.close()
        return jsonify({"error": "Cannot print a voided check"}), 400

    location = mc["location"] or "chatham"
    config = conn.execute("SELECT * FROM check_config WHERE location = ?", (location,)).fetchone()
    if not config:
        config = conn.execute("SELECT * FROM check_config ORDER BY id LIMIT 1").fetchone()
    if not config:
        conn.close()
        return jsonify({"error": "Check config not set up"}), 400

    # Assign check number if not set
    check_num = mc["check_number"]
    if not check_num:
        check_num = str(config["check_number_next"] or 1001)
        conn.execute("UPDATE manual_checks SET check_number = ?, printed_at = datetime('now') WHERE id = ?",
                     (check_num, check_id))
        conn.execute("UPDATE check_config SET check_number_next = ? WHERE location = ?",
                     ((config["check_number_next"] or 1001) + 1, location))
        conn.commit()
    elif not mc["printed_at"]:
        conn.execute("UPDATE manual_checks SET printed_at = datetime('now') WHERE id = ?", (check_id,))
        conn.commit()

    conn.close()

    # Build a payment-like dict for generate_check_pdf
    payment = {
        "vendor_name": mc["payee_name"],
        "amount": mc["amount"],
        "payment_date": (mc["created_at"] or "")[:10],
        "memo": mc["memo"] or "",
    }

    # Build vendor_info from the manual check address fields
    vendor_info = None
    if mc["payee_address_1"]:
        vendor_info = {
            "payment_recipient": mc["payee_name"],
            "remit_address_1": mc["payee_address_1"],
            "remit_address_2": mc["payee_address_2"],
            "remit_city": mc["payee_city"],
            "remit_state": mc["payee_state"],
            "remit_zip": mc["payee_zip"],
        }

    output_path = f"/tmp/manual_check_{check_id}_{check_num}.pdf"
    generate_check_pdf(
        payment=payment,
        invoices=[],
        config=dict(config),
        vendor_info=vendor_info,
        check_number=check_num,
        output_path=output_path,
    )

    return send_file(output_path, mimetype="application/pdf",
                     download_name=f"check_{check_num}_{mc['payee_name']}.pdf", as_attachment=False)


@billpay_bp.route("/api/billpay/manual-checks/<int:check_id>/void", methods=["PUT"])
@admin_required
def void_manual_check(check_id):
    """Void a manual check."""
    conn = get_connection()
    conn.execute("UPDATE manual_checks SET voided = 1, voided_at = datetime('now') WHERE id = ?",
                 (check_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


# ─────────────────────────────────────────────
#  CHECK REGISTER EXPORT (Excel for bookkeeper)
# ─────────────────────────────────────────────

@billpay_bp.route("/api/billpay/check-register/export")
@admin_required
def export_check_register():
    """Export unified check register as Excel — all check types in one sheet."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    date_from = request.args.get("date_from")
    date_to = request.args.get("date_to")
    location = request.args.get("location")

    conn = get_connection()

    # Gather all checks from 3 sources into one list
    checks = []

    # 1. AP vendor checks
    where_ap = ["ap.status != 'voided'"]
    params_ap = []
    if date_from:
        where_ap.append("ap.payment_date >= ?")
        params_ap.append(date_from)
    if date_to:
        where_ap.append("ap.payment_date <= ?")
        params_ap.append(date_to)

    ap_rows = conn.execute(f"""
        SELECT ap.check_number, ap.payment_date as check_date, ap.vendor_name as payee,
               ap.amount, ap.memo, 'Vendor' as check_type, ap.status
        FROM ap_payments ap
        WHERE ap.check_number IS NOT NULL AND ap.check_number != ''
          AND {' AND '.join(where_ap)}
        ORDER BY CAST(ap.check_number AS INTEGER)
    """, params_ap).fetchall()
    for r in ap_rows:
        checks.append(dict(r))

    # 2. Payroll checks
    where_pr = ["pr.voided = 0", "pr.check_number IS NOT NULL"]
    params_pr = []
    if date_from:
        where_pr.append("date(pr.printed_at) >= ?")
        params_pr.append(date_from)
    if date_to:
        where_pr.append("date(pr.printed_at) <= ?")
        params_pr.append(date_to)
    if location:
        where_pr.append("pr.location = ?")
        params_pr.append(location)

    pr_rows = conn.execute(f"""
        SELECT pr.check_number, date(pr.printed_at) as check_date,
               pr.employee_name as payee, pr.net_pay as amount,
               pr.memo, 'Payroll' as check_type, 'printed' as status
        FROM payroll_checks pr
        WHERE {' AND '.join(where_pr)}
        ORDER BY CAST(pr.check_number AS INTEGER)
    """, params_pr).fetchall()
    for r in pr_rows:
        checks.append(dict(r))

    # 3. Manual checks
    where_mc = ["mc.voided = 0", "mc.check_number IS NOT NULL"]
    params_mc = []
    if date_from:
        where_mc.append("date(mc.printed_at) >= ?")
        params_mc.append(date_from)
    if date_to:
        where_mc.append("date(mc.printed_at) <= ?")
        params_mc.append(date_to)
    if location:
        where_mc.append("mc.location = ?")
        params_mc.append(location)

    mc_rows = conn.execute(f"""
        SELECT mc.check_number, date(mc.printed_at) as check_date,
               mc.payee_name as payee, mc.amount,
               mc.memo, mc.check_type, 'printed' as status
        FROM manual_checks mc
        WHERE {' AND '.join(where_mc)}
        ORDER BY CAST(mc.check_number AS INTEGER)
    """, params_mc).fetchall()
    for r in mc_rows:
        checks.append(dict(r))

    # 4. Recurring bill payments (status='paid', has check_number)
    where_rb = ["rbp.status = 'paid'", "rbp.check_number IS NOT NULL", "rbp.check_number != ''"]
    params_rb = []
    if date_from:
        where_rb.append("rbp.paid_date >= ?")
        params_rb.append(date_from)
    if date_to:
        where_rb.append("rbp.paid_date <= ?")
        params_rb.append(date_to)
    if location:
        where_rb.append("(rb.location = ? OR rb.location = 'both')")
        params_rb.append(location)

    rb_rows = conn.execute(f"""
        SELECT rbp.check_number, rbp.paid_date as check_date,
               COALESCE(rb.payable_to, rb.vendor_name) as payee,
               rbp.amount_paid as amount,
               rbp.memo, 'Recurring' as check_type, 'printed' as status
        FROM recurring_bill_payments rbp
        JOIN recurring_bills rb ON rb.id = rbp.bill_id
        WHERE {' AND '.join(where_rb)}
        ORDER BY CAST(rbp.check_number AS INTEGER)
    """, params_rb).fetchall()
    for r in rb_rows:
        checks.append(dict(r))

    conn.close()

    # Sort all checks by check number
    checks.sort(key=lambda x: int(x.get("check_number") or 0))

    # Build Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Check Register"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Title row
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "Red Buoy Public House, Inc. — Check Register"
    title_cell.font = Font(bold=True, size=14)
    if date_from or date_to:
        ws.merge_cells("A2:F2")
        ws["A2"].value = f"Period: {date_from or 'Start'} to {date_to or 'Present'}"
        ws["A2"].font = Font(italic=True, size=10)

    # Column headers
    headers = ["Check #", "Date", "Payee", "Amount", "Type", "Memo"]
    header_row = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    # Data rows
    money_fmt = '#,##0.00'
    total = 0
    for i, ck in enumerate(checks):
        row = header_row + 1 + i
        ws.cell(row=row, column=1, value=int(ck["check_number"])).border = thin_border
        ws.cell(row=row, column=2, value=ck.get("check_date", "")).border = thin_border
        ws.cell(row=row, column=3, value=ck.get("payee", "")).border = thin_border
        amt_cell = ws.cell(row=row, column=4, value=float(ck.get("amount", 0)))
        amt_cell.number_format = money_fmt
        amt_cell.border = thin_border
        ctype = ck.get("check_type", "").title()
        ws.cell(row=row, column=5, value=ctype).border = thin_border
        ws.cell(row=row, column=6, value=ck.get("memo", "")).border = thin_border
        total += float(ck.get("amount", 0))

    # Total row
    total_row = header_row + 1 + len(checks)
    ws.cell(row=total_row, column=3, value="TOTAL").font = Font(bold=True)
    tot_cell = ws.cell(row=total_row, column=4, value=total)
    tot_cell.font = Font(bold=True)
    tot_cell.number_format = money_fmt
    tot_cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    ws.column_dimensions['C'].width = 35
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 40

    # Save to buffer
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"check_register_{date.today().isoformat()}.xlsx"
    return send_file(buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     download_name=fname, as_attachment=True)


# ─────────────────────────────────────────────
#  RECURRING BILLS
# ─────────────────────────────────────────────

def init_recurring_tables():
    """Create recurring_bills, recurring_bill_lines, and recurring_bill_payments tables."""
    conn = get_connection()
    conn.execute("""
        CREATE TABLE IF NOT EXISTS recurring_bills (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vendor_name TEXT NOT NULL,
            description TEXT,
            amount REAL NOT NULL DEFAULT 0,
            frequency TEXT NOT NULL DEFAULT 'monthly',
            start_date TEXT,
            due_day INTEGER DEFAULT 1,
            days_before_due INTEGER DEFAULT 0,
            payment_method TEXT DEFAULT 'check',
            payable_to TEXT,
            location TEXT DEFAULT 'dennis',
            active INTEGER DEFAULT 1,
            created_at TEXT DEFAULT (datetime('now')),
            updated_at TEXT DEFAULT (datetime('now'))
        )
    """)
    # Add days_before_due column if upgrading from earlier version
    try:
        conn.execute("ALTER TABLE recurring_bills ADD COLUMN days_before_due INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        pass
    # Add auto_print column — flagged bills print themselves through the
    # print agent when due (daily 6 PM run in daily_auto_pay_summary.py)
    try:
        conn.execute("ALTER TABLE recurring_bills ADD COLUMN auto_print INTEGER DEFAULT 0")
        conn.commit()
    except Exception:
        pass
    conn.execute("""
        CREATE TABLE IF NOT EXISTS recurring_bill_lines (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_id INTEGER NOT NULL,
            description TEXT NOT NULL,
            amount REAL NOT NULL DEFAULT 0,
            account TEXT,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY(bill_id) REFERENCES recurring_bills(id)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS recurring_bill_payments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bill_id INTEGER NOT NULL,
            due_date TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'paid',
            paid_date TEXT,
            skipped_date TEXT,
            amount_paid REAL,
            check_number TEXT,
            payment_method TEXT,
            memo TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY(bill_id) REFERENCES recurring_bills(id)
        )
    """)
    conn.commit()
    conn.close()


def _get_lines(conn, bill_id):
    rows = conn.execute(
        "SELECT id, description, amount, account, sort_order FROM recurring_bill_lines WHERE bill_id=? ORDER BY sort_order, id",
        (bill_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def _save_lines(conn, bill_id, lines):
    """Replace all lines for a bill and return the total."""
    conn.execute("DELETE FROM recurring_bill_lines WHERE bill_id=?", (bill_id,))
    total = 0.0
    for i, line in enumerate(lines):
        amt = float(line.get("amount") or 0)
        total += amt
        conn.execute(
            "INSERT INTO recurring_bill_lines (bill_id, description, amount, account, sort_order) VALUES (?,?,?,?,?)",
            (bill_id, line.get("description",""), amt, line.get("account",""), i)
        )
    return total


def _next_due_date(frequency, start_date, due_day):
    """Calculate the next due date for a recurring bill from today."""
    from datetime import date as ddate, timedelta
    import calendar as _cal
    today = ddate.today()
    try:
        sd = ddate.fromisoformat(start_date) if start_date else today
    except ValueError:
        sd = today

    due_day = int(due_day or 1)

    if frequency == 'weekly':
        d = sd
        while d <= today:
            d += timedelta(weeks=1)
        return d.isoformat()

    if frequency == 'biweekly':
        d = sd
        while d <= today:
            d += timedelta(weeks=2)
        return d.isoformat()

    if frequency == 'monthly':
        yr, mo = today.year, today.month
        # Try this month's due day
        last_day = _cal.monthrange(yr, mo)[1]
        d = ddate(yr, mo, min(due_day, last_day))
        if d <= today:
            # Move to next month
            mo += 1
            if mo > 12:
                mo = 1; yr += 1
            last_day = _cal.monthrange(yr, mo)[1]
            d = ddate(yr, mo, min(due_day, last_day))
        return d.isoformat()

    if frequency == 'quarterly':
        d = sd
        while d <= today:
            mo = d.month + 3
            yr = d.year + (mo - 1) // 12
            mo = (mo - 1) % 12 + 1
            import calendar as _cal2
            last_day = _cal2.monthrange(yr, mo)[1]
            d = ddate(yr, mo, min(due_day, last_day))
        return d.isoformat()

    if frequency == 'semiannual':
        d = sd
        while d <= today:
            mo = d.month + 6
            yr = d.year + (mo - 1) // 12
            mo = (mo - 1) % 12 + 1
            import calendar as _cal3
            last_day = _cal3.monthrange(yr, mo)[1]
            d = ddate(yr, mo, min(due_day, last_day))
        return d.isoformat()

    if frequency == 'annual':
        d = sd
        while d <= today:
            d = ddate(d.year + 1, d.month, d.day)
        return d.isoformat()

    return today.isoformat()


def _is_due_on(bill, target_date_str):
    """Return True if bill is due on or before target_date and not yet paid/skipped for that period."""
    from datetime import date as ddate, timedelta
    import calendar as _cal
    try:
        target = ddate.fromisoformat(target_date_str)
    except ValueError:
        return False

    freq = bill['frequency']
    start_str = bill['start_date']
    due_day = int(bill['due_day'] or 1)

    try:
        start = ddate.fromisoformat(start_str) if start_str else ddate(target.year, target.month, 1)
    except ValueError:
        start = ddate(target.year, target.month, 1)

    if target < start:
        return False

    # Enumerate all due dates from start up to target
    due_dates = []
    d = start
    while d <= target:
        due_dates.append(d)
        if freq == 'weekly':
            d += timedelta(weeks=1)
        elif freq == 'biweekly':
            d += timedelta(weeks=2)
        elif freq == 'monthly':
            mo = d.month + 1
            yr = d.year + (mo - 1) // 12
            mo = (mo - 1) % 12 + 1
            last_day = _cal.monthrange(yr, mo)[1]
            d = ddate(yr, mo, min(due_day, last_day))
        elif freq == 'quarterly':
            mo = d.month + 3
            yr = d.year + (mo - 1) // 12
            mo = (mo - 1) % 12 + 1
            last_day = _cal.monthrange(yr, mo)[1]
            d = ddate(yr, mo, min(due_day, last_day))
        elif freq == 'semiannual':
            mo = d.month + 6
            yr = d.year + (mo - 1) // 12
            mo = (mo - 1) % 12 + 1
            last_day = _cal.monthrange(yr, mo)[1]
            d = ddate(yr, mo, min(due_day, last_day))
        elif freq == 'annual':
            d = ddate(d.year + 1, d.month, d.day)
        else:
            break

    if not due_dates:
        return False

    # Use the latest due date that is <= target
    candidate = max(dd for dd in due_dates if dd <= target)
    return candidate.isoformat()


@billpay_bp.route("/api/billpay/recurring", methods=["GET"])
@admin_or_accountant_required
def get_recurring_bills():
    include_inactive = request.args.get("include_inactive") == "1"
    conn = get_connection()
    where = [] if include_inactive else ["active = 1"]
    sql = "SELECT * FROM recurring_bills" + (" WHERE " + " AND ".join(where) if where else "") + " ORDER BY vendor_name"
    rows = conn.execute(sql).fetchall()
    result = []
    for r in rows:
        b = dict(r)
        b['next_due_date'] = _next_due_date(b['frequency'], b['start_date'], b['due_day'])
        b['lines'] = _get_lines(conn, b['id'])
        if b['lines']:
            b['amount'] = sum(l['amount'] for l in b['lines'])
        result.append(b)
    conn.close()
    return jsonify(result)


@billpay_bp.route("/api/billpay/recurring/<int:bill_id>", methods=["GET"])
@admin_or_accountant_required
def get_recurring_bill(bill_id):
    conn = get_connection()
    row = conn.execute("SELECT * FROM recurring_bills WHERE id = ?", (bill_id,)).fetchone()
    if not row:
        conn.close()
        return jsonify({"error": "Not found"}), 404
    b = dict(row)
    b['next_due_date'] = _next_due_date(b['frequency'], b['start_date'], b['due_day'])
    b['lines'] = _get_lines(conn, b['id'])
    if b['lines']:
        b['amount'] = sum(l['amount'] for l in b['lines'])
    conn.close()
    return jsonify(b)


@billpay_bp.route("/api/billpay/recurring", methods=["POST"])
@admin_or_accountant_required
def create_recurring_bill():
    data = request.get_json(silent=True) or {}
    vendor_name = data.get("vendor_name", "").strip()
    lines = data.get("lines", [])
    if not vendor_name:
        return jsonify({"error": "vendor_name required"}), 400
    conn = get_connection()
    cur = conn.execute("""
        INSERT INTO recurring_bills (vendor_name, description, amount, frequency,
            start_date, due_day, days_before_due, payment_method, payable_to, location, active, auto_print)
        VALUES (?, ?, 0, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    """, (vendor_name, data.get("description",""),
          data.get("frequency","monthly"), data.get("start_date"),
          int(data.get("due_day") or 1), int(data.get("days_before_due") or 0),
          data.get("payment_method","check"), data.get("payable_to",""),
          data.get("location","dennis"), int(data.get("active", 1)),
          int(data.get("auto_print") or 0)))
    new_id = cur.lastrowid
    total = _save_lines(conn, new_id, lines)
    if total > 0:
        conn.execute("UPDATE recurring_bills SET amount=? WHERE id=?", (total, new_id))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok", "id": new_id}), 201


@billpay_bp.route("/api/billpay/recurring/<int:bill_id>", methods=["PUT"])
@admin_or_accountant_required
def update_recurring_bill(bill_id):
    data = request.get_json(silent=True) or {}
    lines = data.get("lines", [])
    conn = get_connection()
    conn.execute("""
        UPDATE recurring_bills SET vendor_name=?, description=?, frequency=?,
            start_date=?, due_day=?, days_before_due=?, payment_method=?, payable_to=?,
            location=?, active=?, auto_print=?, updated_at=datetime('now')
        WHERE id=?
    """, (data.get("vendor_name",""), data.get("description",""),
          data.get("frequency","monthly"), data.get("start_date"),
          int(data.get("due_day") or 1), int(data.get("days_before_due") or 0),
          data.get("payment_method","check"), data.get("payable_to",""),
          data.get("location","dennis"), int(data.get("active",1)),
          int(data.get("auto_print") or 0), bill_id))
    total = _save_lines(conn, bill_id, lines)
    if total > 0:
        conn.execute("UPDATE recurring_bills SET amount=? WHERE id=?", (total, bill_id))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


@billpay_bp.route("/api/billpay/recurring/<int:bill_id>", methods=["DELETE"])
@admin_required
def delete_recurring_bill(bill_id):
    conn = get_connection()
    conn.execute("DELETE FROM recurring_bill_payments WHERE bill_id = ?", (bill_id,))
    conn.execute("DELETE FROM recurring_bill_lines WHERE bill_id = ?", (bill_id,))
    conn.execute("DELETE FROM recurring_bills WHERE id = ?", (bill_id,))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"})


@billpay_bp.route("/api/billpay/recurring/due", methods=["GET"])
@admin_or_accountant_required
def get_recurring_due():
    """Return bills due within a date window that haven't been paid/skipped.
    ?date=YYYY-MM-DD (default today) ?days=N (default 14) ?location=...
    Respects days_before_due per bill."""
    from datetime import date as ddate, timedelta
    try:
        start = ddate.fromisoformat(request.args.get("date") or date.today().isoformat())
    except ValueError:
        start = date.today()
    window = int(request.args.get("days") or 14)
    location = request.args.get("location")

    conn = get_connection()
    where = ["active = 1"]
    params = []
    if location:
        where.append("(location = ? OR location = 'both')")
        params.append(location)
    rows = conn.execute(
        "SELECT * FROM recurring_bills WHERE " + " AND ".join(where), params
    ).fetchall()

    # Collect unique (bill, due_date) pairs across the window
    seen = set()
    result = []
    for r in rows:
        b = dict(r)
        days_early = int(b.get('days_before_due') or 0)
        # Walk each day in the window
        for offset in range(window + 1):
            check_day = start + timedelta(days=offset)
            effective = (check_day + timedelta(days=days_early)).isoformat()
            due_date_str = _is_due_on(b, effective)
            if not due_date_str:
                continue
            key = (b['id'], due_date_str)
            if key in seen:
                continue
            seen.add(key)
            existing = conn.execute(
                "SELECT id FROM recurring_bill_payments WHERE bill_id=? AND due_date=?",
                (b['id'], due_date_str)
            ).fetchone()
            if existing:
                continue
            entry = dict(b)
            entry['due_date'] = due_date_str
            entry['lines'] = _get_lines(conn, b['id'])
            if entry['lines']:
                entry['amount'] = sum(l['amount'] for l in entry['lines'])
            result.append(entry)

    # Sort by due date
    result.sort(key=lambda x: x['due_date'])
    conn.close()
    return jsonify(result)


@billpay_bp.route("/api/billpay/recurring/<int:bill_id>/pay", methods=["POST"])
@admin_or_accountant_required
def pay_recurring_bill(bill_id):
    data = request.get_json(silent=True) or {}
    paid_date = data.get("paid_date") or date.today().isoformat()
    amount_paid = float(data.get("amount_paid") or 0)
    check_number = data.get("check_number", "")
    payment_method = data.get("payment_method", "check")
    memo = data.get("memo", "")

    conn = get_connection()
    bill = conn.execute("SELECT * FROM recurring_bills WHERE id = ?", (bill_id,)).fetchone()
    if not bill:
        conn.close()
        return jsonify({"error": "Bill not found"}), 404

    due_date = _is_due_on(dict(bill), paid_date)
    if not due_date:
        due_date = paid_date

    # Check for duplicate
    dup = conn.execute(
        "SELECT id FROM recurring_bill_payments WHERE bill_id=? AND due_date=?",
        (bill_id, due_date)
    ).fetchone()
    if dup:
        conn.close()
        return jsonify({"error": "Already recorded for this period"}), 400

    conn.execute("""
        INSERT INTO recurring_bill_payments
            (bill_id, due_date, status, paid_date, amount_paid, check_number, payment_method, memo)
        VALUES (?, ?, 'paid', ?, ?, ?, ?, ?)
    """, (bill_id, due_date, paid_date, amount_paid, check_number, payment_method, memo))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"}), 201


@billpay_bp.route("/api/billpay/recurring/<int:bill_id>/skip", methods=["POST"])
@admin_or_accountant_required
def skip_recurring_bill(bill_id):
    data = request.get_json(silent=True) or {}
    skipped_date = data.get("skipped_date") or date.today().isoformat()

    conn = get_connection()
    bill = conn.execute("SELECT * FROM recurring_bills WHERE id = ?", (bill_id,)).fetchone()
    if not bill:
        conn.close()
        return jsonify({"error": "Bill not found"}), 404

    due_date = _is_due_on(dict(bill), skipped_date)
    if not due_date:
        due_date = skipped_date

    dup = conn.execute(
        "SELECT id FROM recurring_bill_payments WHERE bill_id=? AND due_date=?",
        (bill_id, due_date)
    ).fetchone()
    if dup:
        conn.close()
        return jsonify({"error": "Already recorded for this period"}), 400

    conn.execute("""
        INSERT INTO recurring_bill_payments
            (bill_id, due_date, status, skipped_date)
        VALUES (?, ?, 'skipped', ?)
    """, (bill_id, due_date, skipped_date))
    conn.commit()
    conn.close()
    return jsonify({"status": "ok"}), 201


@billpay_bp.route("/api/billpay/recurring/payments", methods=["GET"])
@admin_or_accountant_required
def get_recurring_payments():
    location = request.args.get("location")
    conn = get_connection()
    where = []
    params = []
    if location:
        where.append("(rb.location = ? OR rb.location = 'both')")
        params.append(location)
    sql = """
        SELECT rbp.id, rb.vendor_name, rb.description, rbp.due_date, rbp.status,
               rbp.paid_date, rbp.skipped_date, rbp.amount_paid, rbp.check_number,
               rbp.payment_method, rbp.memo
        FROM recurring_bill_payments rbp
        JOIN recurring_bills rb ON rb.id = rbp.bill_id
    """
    if where:
        sql += " WHERE " + " AND ".join(where)
    sql += " ORDER BY rbp.due_date DESC, rb.vendor_name"
    rows = conn.execute(sql, params).fetchall()
    conn.close()
    return jsonify([dict(r) for r in rows])
